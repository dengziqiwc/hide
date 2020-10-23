from django.shortcuts import render,HttpResponse
# 引入JsonResponse模块
from django.http import JsonResponse
# Create your views here.
import pymysql
import xlwt
from io import BytesIO
from django.shortcuts import HttpResponse
# 引入处理Excel模块
import openpyxl
# 导入uuid类
import uuid
import win32com.client as win32
import os

import random
# 导入哈希库
import hashlib
import json
import datetime
# 导入Setting
from django.conf import settings
from app01 import models
from pymysql.cursors import DictCursor
from django.core import serializers


def changtime(oldtime):

    oldtime = oldtime.rsplit('.',maxsplit=1)[0]
    newtime = oldtime.replace('T',' ')
    utc_date = datetime.datetime.strptime(newtime, "%Y-%m-%d %H:%M:%S")
    local_date = utc_date + datetime.timedelta(hours=8)
    local_date_str = datetime.datetime.strftime(local_date, '%Y-%m-%d %H:%M:%S')
    # print(local_date_str)  # 2019-07-26 16:20:54

    dateTime_p = datetime.datetime.strptime(local_date_str, '%Y-%m-%d %H:%M:%S')
    print(dateTime_p)  # 2019-01-30 15:29:08
    return dateTime_p


def get_random_str():
    #获取uuid的随机数
    uuid_val = uuid.uuid4()
    #获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')
    #获取md5实例
    md5 = hashlib.md5()
    #拿取uuid的md5摘要
    md5.update(uuid_str)
    #返回固定长度的字符串
    return md5.hexdigest()

#获取三元素的随机数
def generate_random_str(randomlength=16):
    """
    生成一个指定长度的随机字符串
    """
    random_str = ''
    base_str = 'ABCDEFGHIGKLMNOPQRSTUVWXYZabcdefghigklmnopqrstuvwxyz0123456789'
    length = len(base_str) - 1
    for i in range(randomlength):
        random_str += base_str[random.randint(0, length)]
    return random_str

def export_xls_out(request):
    conn =pymysql.connect(host='127.0.0.1',user='root',passwd='123456',db='day11',port=3306,charset='utf8',cursorclass=pymysql.cursors.DictCursor)
#数据库参数自己配置
    sql='select * from app01_device'
    cur = conn.cursor()
    cur.execute(sql)
    result = cur.fetchall()#获取查询结果
    print(result)
    response = HttpResponse(content_type='application/vnd.ms-excel')#指定返回为excel文件
    response['Content-Disposition'] = 'attachment;filename=test.xls'#指定返回文件名
    wb = xlwt.Workbook(encoding = 'utf-8')#设定编码类型为utf8
    sheet = wb.add_sheet(u'表格')#excel里添加类别
    sheet.write(0,0,'id')
    sheet.write(0,1,'devicename')
    sheet.write(0,2,'devicesecret')
    sheet.write(0,3,'productname')

    row = 1
    for list in result:
        sheet.write(row,0, list['id'])
        sheet.write(row,1, list['devicename'])
        sheet.write(row,2, list['devicesecret'])
        sheet.write(row,3, list['productname_id'])
        row=row + 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response


def page(request):
    return render(request,'page.html')



def accountcreate(request):
    data = request.body.decode('utf-8')
    data = eval(data)
    print(data,type(data))
    print(data['productname'])
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    productkey = generate_random_str(11)
    productsecret = generate_random_str(16)
    sql = "INSERT INTO app01_product(productname,productkey,productsecret) VALUES ({},{},{})".format(data['productname'],productkey,productsecret)
    print(sql)
    # cursor.execute(sql)
    return JsonResponse({'code':1,'data':123})
#增加产品
def productcreate(request):
    data = request.body.decode('utf-8')
    data = eval(data)
    print(data,type(data))
    print(data['productname'])
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    productkey = generate_random_str(11)
    productsecret = generate_random_str(16)
    sql = "INSERT INTO app01_product(productname,productkey,productsecret) VALUES('{}','{}','{}')".format(data['productname'],productkey,productsecret)
    print(sql)
    cursor.execute(sql)
    conn.commit()
    return JsonResponse({'code':1,'data':123})
#查到所有的产品
def productall(request):
    conn = pymysql.connect(host='192.168.1.114',port=3306,user='root',passwd='654321',db='day3',charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = 'SELECT productname FROM app01_product'
    cursor.execute(sql)
    data = cursor.fetchall()
    print('cc',data)
    return JsonResponse({'code':1,'data':data})

#创建设备
def devicecreate(request):
    data = request.body.decode('utf-8')
    print(data)
    data = eval(data)
    devicename = data.get('devicename')
    productname = data.get('productname')
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    devicesecret = generate_random_str(30)
    sql = "INSERT INTO app01_device(devicename,devicesecret,productname_id) VALUES('{}','{}','{}')".format(devicename,devicesecret,productname)
    print(sql)
    col = cursor.execute(sql)
    print(col)
    conn.commit()
    return JsonResponse({'code':1,'data':11})

def export_student_excel(request,typee):
    """到处数据到excel"""
    print('类型',typee)
    # 获取信息
    if typee=='humidity':
        table_name = 'app01_hh'
    else:
        table_name = 'app01_tt'
    data = request.body.decode('utf-8')
    data = eval(data)
    print(data, type(data))
    # pid = data.get('id')
    mform = data.get('msg')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    end_time = changtime(end_time)
    start_time = changtime(start_time)

    # print(start_time, end_time)
    # print(type(start_time), type(end_time))
    # objs = models.Test2.objects.filter(event_date__range=(start_time, end_time)).values()
    #连接数据库
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "select * from {} where t0_time_before<'{}'and t0_time_before>'{}'LIMIT 1000".format(table_name,end_time, start_time)
    print(sql)
    cursor.execute(sql)
    res = cursor.fetchall()
    # 准备名称
    # print('cccccccc',mform,type(mform))
    pro_name = get_random_str()
    excel_name = pro_name + ".xlsx"
    # 准备写入的路劲
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入到Excel
    print(1)
    write_to_excel(res, path,mform)
    print(2)
    if mform=='xls':
        excel_name = pro_name+'.xls'
    print('最新的excelname',excel_name)
    # 返回
    return JsonResponse({'code':1, 'name':excel_name })

#导出多个excel
def export_xls_many(request):
    """导出数据到excel"""
    data = request.body.decode('utf-8')
    data = eval(data)
    print(data)
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    end_time = changtime(end_time)
    start_time = changtime(start_time)
    # conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
    conn = pymysql.connect(host='192.168.1.114',port=3306,user='root',passwd='654321',db='day1',charset='utf8')
    cursor = conn.cursor(DictCursor)
    s = ['app01_tt','app01_hh']
    excel_name_list = []
    for i in s:
        sql = 'select * from {} LIMIT 1000'.format(i)
        cursor.execute(sql)
        res = cursor.fetchall()
        print(res)
        print(start_time,end_time)
        # 准备名称
        excel_name = get_random_str() + ".xlsx"
        excel_name_list.append(excel_name)
        # 准备写入的路劲
        path = os.path.join(settings.MEDIA_ROOT, excel_name)
        # 写入到Excel
        write_to_excel(res, path)
    # print(111111111111)
    return JsonResponse({'code': 1, 'name': excel_name_list})
    # 获取信息
    # data = request.body.decode('utf-8')
    # data = eval(data)
    # print(data, type(data))
    # pid = data.get('id')
    # start_time = data.get('start_time')
    # end_time = data.get('end_time')
    # end_time = changtime(end_time)
    # start_time = changtime(start_time)
    # # 连接数据库
    # conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    # cursor = conn.cursor(DictCursor)
    # sql = "select * from app01_hh where t0_time_before<'{}'and t0_time_before>'{}'".format(end_time, start_time)
    # cursor.execute(sql)
    # res = cursor.fetchmany(10)
    # # 准备名称
    # excel_name = get_random_str() + ".xlsx"
    # # 准备写入的路劲
    # path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # # 写入到Excel
    # write_to_excel(res, path)
    # # 返回
    # return JsonResponse({'code': 1, 'name': excel_name})

import pythoncom
def transform(oldpath):
    # fileList=os.listdir(parent_path)  #文件夹下面所有的文件
    # num=len(fileList)
    # for i in range(num):
    #     file_Name=os.path.splitext(fileList[i])   #文件和格式分开
    #     if file_Name[1]=='.xlsx':
    #         tranfile1=parent_path+'\\'+fileList[i]  #要转换的excel
    #         tranfile2=out_path+'\\'+file_Name[0]    #转换出来excel
    #         excel=win32.gencache.EnsureDispatch('excel.application')
    #         pro=excel.Workbooks.Open(tranfile1)   #打开要转换的excel
    #         pro.SaveAs(tranfile2+".xls",FileFormat=56)  #另存为xls格式
    #         pro.Close()
    #         excel.Application.Quit()
    print('1111',oldpath)
    file_Name = os.path.splitext(oldpath)  # 文件和格式分开
    print(file_Name[0])
    pathy = file_Name[0] + '.xls'
    print(pathy)
    pythoncom.CoInitialize()
    excel = win32.gencache.EnsureDispatch('excel.application')
    pro = excel.Workbooks.Open(oldpath)  # 打开要转换的excel
    pro.SaveAs(pathy, FileFormat=56)  # 另存为xls格式


def write_to_excel(data:list, path:str,mform='xlsx'):
    """把数据库写入到Excel"""
    # print(mform,path,data)
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'student'
    # 准备keys
    keys = data[0].keys()
    # print('kkkkkk',keys)
    # sheet.cell()
    # 准备写入数据
    s=1
    for i in keys:
        sheet.cell(row=1,column=s,value=i)
        s=s+1
    for index, item in enumerate(data):
        # 遍历每一个元素
        # print('iiiii',index,item)
        for k,v in enumerate(keys):
            # print('ttttt',k,v)
            sheet.cell(row=index + 2, column=k+ 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)
    print('2222',path)
    if mform=='xls':
        transform(path)

def query_devices(request):
    def query_students(request):
        """查询设备信息"""
        # 接收传递过来的查询条件--- axios默认是json --- 字典类型（'inputstr'）-- data['inputstr']
        data = json.loads(request.body.decode('utf-8'))
        try:
            # 使用ORM获取满足条件的学生信息 并把对象转为字典格式
            obj_devices = models.TT.objects.values()
            # 把外层的容器转为List
            devices = list(obj_devices)
            # 返回
            return JsonResponse({'code': 1, 'data': devices})
        except Exception as e:
            # 如果出现异常，返回
            return JsonResponse({'code': 0, 'msg': "查询学生信息出现异常，具体错误：" + str(e)})

def deviceshow(request):
    # device_data = models.Device.objects.values_list('devicename','devicesecret','productname')
    # return render(request,'deviceshow.html',{'device_data':device_data})
    """获取所有设备的信息"""
    try:
        conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
        cursor = conn.cursor(DictCursor)
        sql = "select devicename as'设备名',devicesecret as '设备密钥',productname_id as'产品名'  from app01_device"

        cursor.execute(sql)
        devices = cursor.fetchall()
        # device_data = models.Device.objects.values('pk','devicename','devicesecret','productname')
        # # 把外层的容器转为List
        # devices= list(device_data)
        # 返回
        return JsonResponse({'code':1, 'data':devices})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取设备信息出现异常，具体错误：" + str(e)})

def productshow(request):
    # product_data = models.Product.objects.values('productname','productkey','productsecret')
    # # 把外层的容器转为List
    # products = list(product_data)
    """获取所有产品的信息"""

    try:
        conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day3', port=3306, charset='utf8')
        cursor = conn.cursor(DictCursor)
        sql = "select productname as'产品名',productkey as '产品键值',productsecret as'产品密钥'  from app01_product"
        cursor.execute(sql)
        products = cursor.fetchall()
        # product_data = models.Product.objects.values('productname', 'productkey', 'productsecret')
        # # 把外层的容器转为List
        # products = list(product_data)
        # 返回
        print(products)
        return JsonResponse({'code':1, 'data':products})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取设产品信息出现异常，具体错误：" + str(e)})

# UTC 转换为 格式化的时间字符串
# def u2s(origin_date_str):
#     utc_date = datetime.datetime.strptime(origin_date_str, "%Y-%m-%d %H:%M:%S")
#     local_date = utc_date + datetime.timedelta(hours=8)
#     local_date_str = datetime.datetime.strftime(local_date, '%Y-%m-%d %H:%M:%S')
#     print(local_date_str)  # 2019-07-26 16:20:54
#     return local_date_str

#查数据

def querydevicehumidity(request):
    data = request.body.decode('utf-8')
    print(data)
    data = eval(data)
    # print(data,type(data))
    # pid = data.get('id')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    end_time = changtime(end_time)

    start_time = changtime(start_time)

    # print(start_time,end_time)
    # print(type(start_time),type(end_time))

    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "select * from app01_hh where t0_time_before<'{}'and t0_time_before>'{}' LIMIT 1000".format(end_time,start_time)
    # print('sql',sql)
    cursor.execute(sql)
    res = cursor.fetchall()
    # res = cursor.fetchmany(10)
    # res = cursor.fetchone()
    # print(type(res),res)
    # res_list = list(res)
    # objs = models.Test2.objects.filter(event_date__range=(start_time,end_time))
    # print(objs)
    # objs_list = list(objs)
    # print(objs_list)
    # data = {'data':objs_list}
    # objs_list = serializers.serialize('json',objs)


    # data = {'code':111,'data':res_list}
    data = {'code':111,'data':res}
    return JsonResponse(data)
def querydevicetemperature(request):
    data = request.body.decode('utf-8')
    data = eval(data)
    # pid = data.get('id')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    end_time = changtime(end_time)
    start_time = changtime(start_time)
    print(end_time,start_time)
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "select * from app01_tt where t0_time_before<'{}'and t0_time_before>'{}'".format(end_time, start_time)
    print(sql)
    cursor.execute(sql)
    res = cursor.fetchmany(10)
    data = {'code': 111, 'data': res}
    print(data)
    return JsonResponse(data)
    # data = request.body.decode('utf-8')
    # data = eval(data)
    # print(data,type(data))
    # pid = data.get('id')
    # start_time = data.get('start_time')
    # end_time = data.get('end_time')
    # end_time = changtime(end_time)
    # start_time = changtime(start_time)
    #
    # print(start_time,end_time)
    # print(type(start_time),type(end_time))
    # objs = models.Test2.objects.filter(event_date__range=(start_time,end_time))
    # # print(objs)
    # # objs_list = list(objs)
    # # print(objs_list)
    # # data = {'data':objs_list}
    # objs_list = serializers.serialize('json',objs)
    #
    #
    # data = {'data':objs_list}
    # return JsonResponse(data)

#获取当前的湿度
def getcurrenth(request,productname,devicename):
    print('成功了吗',productname,devicename)
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "SELECT * FROM app01_hh ORDER BY id DESC LIMIT 1"
    print(sql)
    cursor.execute(sql)
    res = cursor.fetchone()
    print(res,type(res))
    # obj = models.Test2.objects.last()
    data = {'msg':'湿度','currenth':res['t0_CurrentHumidity'],'currentTime':res['t0_time_before']}
    return JsonResponse(data)
    # return HttpResponse(123)
#获取当前的湿度
def getcurrenthc(request):
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "SELECT * FROM app01_hh ORDER BY id DESC LIMIT 1"
    print(sql)
    cursor.execute(sql)
    res = cursor.fetchone()
    print(res,type(res))
    # obj = models.Test2.objects.last()
    data = {'msg':'湿度','currenth':res['t0_CurrentHumidity'],'currentTime':res['t0_time_before']}
    return JsonResponse(data)
#获取多当前的温度
def getcurrentt(request,productname,devicename):
    print('温度',productname,devicename)
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "SELECT * FROM app01_tt ORDER BY id DESC LIMIT 1"
    cursor.execute(sql)
    res = cursor.fetchone()
    data = {'msg': '温度', 'currentt': res['t0_CurrentTemperature'], 'currentTime': res['t0_time_before']}
    print('数据',data)
    return JsonResponse(data)
#获取当前的温度
def getcurrenttc(request):
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = "SELECT * FROM app01_tt ORDER BY id DESC LIMIT 1"
    cursor.execute(sql)
    res = cursor.fetchone()
    data = {'msg': '温度', 'currentt': res['t0_CurrentTemperature'], 'currentTime': res['t0_time_before']}
    return JsonResponse(data)

#图表测试视图
def charttest(request):
    data = request.body.decode('utf-8')
    print(data)
    data = eval(data)
    print(data,type(data))
    # pid = data.get('id')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    end_time = changtime(end_time)
    start_time = changtime(start_time)

    print(start_time,end_time)
    print(type(start_time),type(end_time))
    print('开始了')
    conn = pymysql.connect(host='192.168.1.114', user='root', passwd='654321', db='day1', port=3306, charset='utf8')
    cursor = conn.cursor(DictCursor)
    # sql = "select * from app01_hh LIMIT 1000 "
    sql = "select * from app01_hh where t0_time_before<'{}'and t0_time_before>'{}' LIMIT 1000".format(end_time, start_time)
    cursor.execute(sql)
    print(sql)
    res = cursor.fetchall()
    data = {'code': 111, 'data': res}
    print(data)
    print('结束了')
    return JsonResponse(data)

#错误信息读取
def dataerror(request):
    conn = pymysql.connect(host='192.168.1.114',port=3306,user='root',passwd='654321',db='day1',charset='utf8')
    cursor = conn.cursor(DictCursor)
    sql = 'SELECT productname,devicename from errordata'
    col = cursor.execute(sql)
    print('cc',col,type(col))
    data = {'code':0,'data':''}
    if col != 0:
        data['data'] = cursor.fetchall()
        data['code'] = 1
        print('cc',data)
        # sql = 'DELETE FROM errordata'
        # cursor.execute(sql)
        # conn.commit()

    return JsonResponse(data)